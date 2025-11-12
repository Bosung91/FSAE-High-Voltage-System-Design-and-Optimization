import os
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font

# === CONFIGURATION ===
TARGET_FOLDER = "FSAE-High-Voltage-System-Design-and-Optimization"
EXCLUDED_FILE = "appendix.md"
OUTPUT_FILE = "word_count_report.xlsx"


def clean_markdown_syntax(text):
    """
    Remove Markdown-specific syntax while keeping visible text.
    """
    # Remove Markdown table lines (| ... |)
    text = re.sub(r'^\|.*\|\s*$', '', text, flags=re.MULTILINE)

    # Remove Markdown images
    text = re.sub(r'!\[.*?\]\(.*?\)', '', text)

    # Remove links but keep their label text
    text = re.sub(r'\[([^\]]+)\]\([^)]+\)', r'\1', text)

    # Remove formatting (*, #, etc.)
    text = re.sub(r'[`*_#>\-~]+', '', text)

    return text


def split_markdown_html(content):
    """
    Separate Markdown text and embedded HTML text from a Markdown file.
    Returns (markdown_text, html_text).
    """
    # Detect HTML tags
    html_parts = re.findall(r'<[^>]+>.*?</[^>]+>', content, flags=re.DOTALL)
    html_text = "\n".join(html_parts)

    # Remove HTML sections to get pure Markdown part
    markdown_only = re.sub(r'<[^>]+>.*?</[^>]+>', '', content, flags=re.DOTALL)

    return markdown_only, html_text


def extract_visible_text_markdown(text):
    """
    Clean and return visible text from Markdown (excluding HTML and tables).
    """
    text = clean_markdown_syntax(text)
    soup = BeautifulSoup(text, "html.parser")

    # Remove table content
    for table in soup.find_all("table"):
        table.decompose()

    # Remove invisible tags
    for tag in soup(["script", "style", "code", "noscript", "meta", "link"]):
        tag.decompose()

    text = soup.get_text(separator=" ", strip=True)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def extract_visible_text_html(text):
    """
    Extract visible text from embedded HTML within Markdown.
    """
    soup = BeautifulSoup(text, "html.parser")

    # Remove tables, scripts, and non-visible elements
    for tag in soup(["table", "script", "style", "noscript", "meta", "link"]):
        tag.decompose()

    html_visible = soup.get_text(separator=" ", strip=True)
    html_visible = re.sub(r'\s+', ' ', html_visible).strip()
    return html_visible


def word_count_like_word(text):
    """
    Count words using Microsoft Word's logic.
    """
    pattern = re.compile(
        r"""
        [A-Za-zÀ-ÖØ-öø-ÿ]+(?:[-'][A-Za-zÀ-ÖØ-öø-ÿ]+)*  # Hyphenated words
        | \d+(?:[.,]\d+)*                              # Numbers with commas/decimals
        | [A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,} # Emails
        | https?://\S+                                  # URLs
        """,
        re.VERBOSE
    )
    return len(pattern.findall(text))


def count_words_in_markdown(folder_path):
    """
    Recursively count words in Markdown files (excluding appendix.md and tables),
    and provide breakdowns for Markdown-only and embedded HTML content.
    """
    total_words = 0
    file_word_data = {}

    for root, _, files in os.walk(folder_path):
        for filename in files:
            if filename.lower().endswith(".md") and filename.lower() != EXCLUDED_FILE:
                path = os.path.join(root, filename)
                try:
                    with open(path, "r", encoding="utf-8") as f:
                        content = f.read()
                except Exception as e:
                    print(f"⚠️ Could not read {path}: {e}")
                    continue

                md_part, html_part = split_markdown_html(content)

                markdown_text = extract_visible_text_markdown(md_part)
                html_text = extract_visible_text_html(html_part)

                md_count = word_count_like_word(markdown_text)
                html_count = word_count_like_word(html_text)
                total_count = md_count + html_count

                relative_path = os.path.relpath(path, folder_path)
                file_word_data[relative_path] = {
                    "markdown_words": md_count,
                    "html_words": html_count,
                    "total_words": total_count,
                }
                total_words += total_count

    return total_words, file_word_data


def save_to_excel(file_word_data, total_words, output_file):
    """
    Save word counts (with Markdown/HTML breakdown) to an Excel file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Word Count Report"

    # Header
    ws.append(["Folder", "File Name", "Markdown Words", "HTML Words", "Total Words"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Data
    for path, data in sorted(file_word_data.items()):
        folder, file_name = os.path.split(path)
        ws.append([folder, file_name, data["markdown_words"], data["html_words"], data["total_words"]])

    # Total row
    ws.append(["", "TOTAL (excluding appendix.md, tables)", "", "", total_words])
    ws.cell(ws.max_row, 2).font = Font(bold=True)
    ws.cell(ws.max_row, 5).font = Font(bold=True)

    # Auto column width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    wb.save(output_file)
    print(f"\n✅ Word count report saved as: {output_file}")


if __name__ == "__main__":
    if not os.path.exists(TARGET_FOLDER):
        print(f"❌ Folder not found: {TARGET_FOLDER}")
    else:
        total, data = count_words_in_markdown(TARGET_FOLDER)

        print("Word count by file:")
        for file, d in sorted(data.items()):
            print(f"  {file}: Markdown={d['markdown_words']} | HTML={d['html_words']} | Total={d['total_words']}")

        print(f"\nTotal (excluding appendix.md and table contents): {total}")
        save_to_excel(data, total, OUTPUT_FILE)
