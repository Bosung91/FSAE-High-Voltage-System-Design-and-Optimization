import re
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font


# ==================================================
# 1. TEXT CLEANING WITH REQUIRED EXCLUSIONS
# ==================================================
EXCLUDE_ANCHOR_TEXTS = [
    "Previous Section",
    "Next Section",
    "List of Abbreviations",
    "FSAE-High-Voltage-System-Design-and-Optimization"
]


def fetch_webpage_text(url, exclude_tables=True):
    """
    Fetch webpage, extract visible text, exclude unwanted elements.
    """
    try:
        response = requests.get(url, timeout=15)
        response.raise_for_status()
    except Exception as e:
        print(f"❌ Error fetching {url}: {e}")
        return ""

    soup = BeautifulSoup(response.text, "html.parser")

    # Remove unwanted elements
    for tag in soup(["script", "style", "noscript", "meta", "link", "header", "footer", "nav", "form"]):
        tag.decompose()

    # REMOVE <a> tags whose text should be excluded
    for a in soup.find_all("a"):
        if a.get_text(strip=True) in EXCLUDE_ANCHOR_TEXTS:
            a.decompose()

    # Remove TABLES (optional)
    if exclude_tables:
        for table in soup.find_all("table"):
            table.decompose()

    # REMOVE FIGURE LABELS (Figure X:, Fig. X, etc.)
    for tag in soup.find_all(text=True):
        cleaned = re.sub(r'\b(Figure|Fig\.?)\s*\d+[:.\-]?\s*.*', '', tag, flags=re.IGNORECASE)
        if cleaned.strip() == "":
            tag.extract()

    # Extract visible text
    text = soup.get_text(separator=" ", strip=True)

    # Normalize whitespace
    text = re.sub(r"\s+", " ", text).strip()

    return text


# ==================================================
# 2. MICROSOFT-WORD STYLE WORD COUNT
# ==================================================
def word_count_like_word(text):
    """
    Count words using Microsoft Word’s word-count logic.
    """
    pattern = re.compile(
        r"""
        [A-Za-zÀ-ÖØ-öø-ÿ]+(?:[-'][A-Za-zÀ-ÖØ-öø-ÿ]+)*      # hyphenated words
        | \d+(?:[.,]\d+)*                                   # numbers
        | [A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}    # emails
        | https?://\S+                                      # URLs
        """,
        re.VERBOSE
    )
    return len(pattern.findall(text))


# ==================================================
# 3. PROCESS MULTIPLE WEBPAGES
# ==================================================
def analyze_webpages(urls, exclude_tables=True):
    results = []
    total = 0

    for url in urls:
        text = fetch_webpage_text(url, exclude_tables)
        count = word_count_like_word(text)
        total += count

        results.append({
            "url": url,
            "characters": len(text),
            "word_count": count
        })

        print(f"🔗 {url}")
        print(f"   Visible text: {len(text):,} characters")
        print(f"   Word count (Word-style): {count:,}\n")

    print(f"✅ Combined total: {total:,} words")
    return results, total


# ==================================================
# 4. EXPORT RESULTS TO EXCEL
# ==================================================
def save_results_to_excel(results, total, output_file="webpage_word_count.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Webpage Word Count"

    ws.append(["URL", "Visible Characters", "Word Count"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in results:
        ws.append([r["url"], r["characters"], r["word_count"]])

    # Total row
    ws.append(["TOTAL", "", total])
    ws.cell(ws.max_row, 1).font = Font(bold=True)
    ws.cell(ws.max_row, 3).font = Font(bold=True)

    # Auto-size columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    wb.save(output_file)
    print(f"\n📊 Results saved to: {output_file}")


# ==================================================
# 5. EXAMPLE USAGE
# ==================================================
if __name__ == "__main__":
    urls = [
        'https://bosung91.github.io/FSAE-High-Voltage-System-Design-and-Optimization/',
        'https://bosung91.github.io/FSAE-High-Voltage-System-Design-and-Optimization/introduction.html',
        'https://bosung91.github.io/FSAE-High-Voltage-System-Design-and-Optimization/objective-and-scope.html',
        'https://bosung91.github.io/FSAE-High-Voltage-System-Design-and-Optimization/context-of-problem.html',
        'https://bosung91.github.io/FSAE-High-Voltage-System-Design-and-Optimization/R25evo/r25evo.html',
        'https://bosung91.github.io/FSAE-High-Voltage-System-Design-and-Optimization/R26e/precharge-discharge-pcb.html',
        'https://bosung91.github.io/FSAE-High-Voltage-System-Design-and-Optimization/R26e/tractive-battery-pdm.html',
        'https://bosung91.github.io/FSAE-High-Voltage-System-Design-and-Optimization/R26e/hv-distribution-pcb.html',
        'https://bosung91.github.io/FSAE-High-Voltage-System-Design-and-Optimization/shortcomings.html',
        'https://bosung91.github.io/FSAE-High-Voltage-System-Design-and-Optimization/future-work.html',
        'https://bosung91.github.io/FSAE-High-Voltage-System-Design-and-Optimization/references.html'
    ]

    results, total = analyze_webpages(urls, exclude_tables=True)
    save_results_to_excel(results, total)
